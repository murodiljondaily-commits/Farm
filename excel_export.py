"""Excel farm report generator — replaces the deprecated Google Sheets export
(sheets_sync.py, removed). Produces a 3-sheet .xlsx as raw bytes, ready to
stream back from a FastAPI endpoint. All labels come from excel_labels.py,
not inline strings here.
"""
from datetime import datetime, timezone
from io import BytesIO
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

import excel_labels as lbl
import firestore_db

HEADER_FONT = Font(bold=True, color="FFFFFFFF")
HEADER_FILL = PatternFill(start_color="FF1F5C52", end_color="FF1F5C52", fill_type="solid")


def _write_header(ws: Worksheet, headers: List[str]) -> None:
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def _autosize(ws: Worksheet, headers: List[str], rows: List[List]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        longest = len(str(header))
        for row in rows:
            val = row[col_idx - 1]
            longest = max(longest, len(str(val)) if val is not None else 0)
        # Uzbek text runs longer than English equivalents — pad generously.
        ws.column_dimensions[get_column_letter(col_idx)].width = min(longest + 4, 45)


def _fmt_date(iso_str: Optional[str]) -> str:
    if not iso_str:
        return ""
    try:
        dt = datetime.fromisoformat(iso_str.replace("Z", "+00:00"))
        return dt.strftime("%d.%m.%Y")
    except Exception:
        return iso_str


def _fill_for(status: str) -> Optional[PatternFill]:
    hex_color = lbl.fill_color_for_status(status)
    if not hex_color:
        return None
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _closed_in_period(closed_at_iso: str, period_start: datetime) -> bool:
    try:
        dt = datetime.fromisoformat(closed_at_iso.replace("Z", "+00:00"))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt >= period_start
    except Exception:
        return False


async def generate_farm_excel(farm_id: str) -> bytes:
    animals = await firestore_db.get_all_animals(farm_id)
    cases = await firestore_db.get_all_cases(farm_id)

    wb = Workbook()

    # ── Sheet 1: Hayvonlar ro'yxati ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = lbl.SHEET_ROSTER
    _write_header(ws1, lbl.ROSTER_HEADERS)
    roster_rows: List[List] = []
    for i, a in enumerate(animals, start=1):
        status = a.get("status", "")
        row = [
            i,
            a.get("ear_tag", ""),
            a.get("name", "") or "",
            a.get("species", ""),
            a.get("breed", "") or "",
            lbl.animal_status_label(status),
            _fmt_date(a.get("updated_at")),
        ]
        roster_rows.append(row)
        ws1.append(row)
        fill = _fill_for(status)
        if fill:
            ws1.cell(row=i + 1, column=6).fill = fill
    _autosize(ws1, lbl.ROSTER_HEADERS, roster_rows)

    # ── Sheet 2: Sog'liq voqealari ────────────────────────────────────────────
    ws2 = wb.create_sheet(lbl.SHEET_HEALTH)
    _write_header(ws2, lbl.HEALTH_HEADERS)
    cases_sorted = sorted(cases, key=lambda c: c.get("opened_at") or "", reverse=True)
    health_rows: List[List] = []
    for i, c in enumerate(cases_sorted, start=1):
        status = c.get("status", "open")
        row = [
            _fmt_date(c.get("opened_at")),
            c.get("ear_tag", "") or "",
            c.get("animal_name", "") or "",
            ", ".join(c.get("symptoms", []) or []) or c.get("visual_findings", "") or "",
            c.get("severity", ""),
            lbl.case_status_label(status),
            lbl.case_outcome_label(c.get("outcome")),
        ]
        health_rows.append(row)
        ws2.append(row)
        fill = _fill_for(status)
        if fill:
            ws2.cell(row=i + 1, column=6).fill = fill
    _autosize(ws2, lbl.HEALTH_HEADERS, health_rows)

    # ── Sheet 3: Statistika ───────────────────────────────────────────────────
    ws3 = wb.create_sheet(lbl.SHEET_STATS)
    ws3.append([lbl.STATS_LABELS["title"]])
    ws3["A1"].font = Font(bold=True, size=14)
    ws3.append([])
    ws3.append([lbl.STATS_LABELS["total_animals"], len(animals)])

    active_cases = [c for c in cases if not c.get("closed_at")]
    severity_counts: Dict[str, int] = {}
    for c in active_cases:
        sev = c.get("severity", "low")
        severity_counts[sev] = severity_counts.get(sev, 0) + 1

    ws3.append([])
    ws3.append([lbl.STATS_LABELS["active_cases_by_severity"]])
    for sev_key in ("low", "medium", "high", "emergency"):
        label = lbl.STATS_LABELS.get(f"severity_{sev_key}", sev_key)
        ws3.append([label, severity_counts.get(sev_key, 0)])

    now = datetime.now(timezone.utc)
    period_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    resolved_this_period = sum(
        1 for c in cases
        if c.get("closed_at") and _closed_in_period(c["closed_at"], period_start)
    )
    ws3.append([])
    ws3.append([lbl.STATS_LABELS["resolved_this_period"], resolved_this_period])

    ws3.column_dimensions["A"].width = 45
    ws3.column_dimensions["B"].width = 20

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
